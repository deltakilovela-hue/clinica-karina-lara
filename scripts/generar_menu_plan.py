#!/usr/bin/env python3
"""
Genera Menu_Semanal_[Paciente].xlsx con datos reales del plan nutricional.
Uso: python3 generar_menu_plan.py <input.json> <output.xlsx>
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime

if len(sys.argv) < 3:
    print("Uso: generar_menu_plan.py <input.json> <output.xlsx>", file=sys.stderr)
    sys.exit(1)

with open(sys.argv[1], encoding='utf-8') as f:
    datos = json.load(f)

output_path = sys.argv[2]
paciente = datos.get('paciente', {}) or {}
menu     = datos.get('menu', {})   # { "LUNES": { "desayuno": "...", ... }, ... }

nombre_pac = paciente.get('nombre', 'Paciente') or 'Paciente'
edad_pac   = str(paciente.get('edad', ''))
semana_actual = datetime.now().strftime("semana del %d/%m/%Y")

# ── Colores ───────────────────────────────────────────────────────────────────
ROJO      = "7B1B2A"
ROJO_CLR  = "F5E8EB"
VERDE     = "2D6A4F"
VERDE_CLR = "D8F3DC"
NARANJA   = "E65100"
NARANJA_C = "FFF3E0"
MORADO    = "4A148C"
MORADO_C  = "F3E5F5"
AZUL      = "1565C0"
AZUL_C    = "E3F2FD"
GRIS_H    = "37474F"
GRIS_CLR  = "ECEFF1"
BLANCO    = "FFFFFF"
NEGRO     = "1A1A1A"
DORADO    = "8B6914"
BG_NOTA   = "FFFDE7"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(c=NEGRO, b=False, sz=9, it=False):
    return Font(name="Arial", color=c, bold=b, size=sz, italic=it)
def thin():  return Side(style="thin",   color="CCCCCC")
def med():   return Side(style="medium", color="999999")
def ba():    return Border(left=thin(), right=thin(), top=thin(), bottom=thin())
def bo():    return Border(left=med(),  right=med(),  top=med(),  bottom=med())
def ca():    return Alignment(horizontal="center", vertical="center",  wrap_text=True)
def la():    return Alignment(horizontal="left",   vertical="top",     wrap_text=True)
def lc():    return Alignment(horizontal="left",   vertical="center",  wrap_text=True)

DIAS = ["LUNES","MARTES","MIÉRCOLES","JUEVES","VIERNES","SÁBADO","DOMINGO"]

COMIDAS = [
    ("🌅 DESAYUNO",     AZUL,    AZUL_C,    "desayuno"),
    ("🍎 COLACIÓN AM",  VERDE,   VERDE_CLR, "colacion_am"),
    ("🍽️ COMIDA",       ROJO,    ROJO_CLR,  "comida"),
    ("🥜 COLACIÓN PM",  NARANJA, NARANJA_C, "colacion_pm"),
    ("🌙 CENA",         MORADO,  MORADO_C,  "cena"),
]

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Menú Semanal"

# Dimensiones de columnas
ws.column_dimensions["A"].width = 15
for col in range(2, 9):
    ws.column_dimensions[get_column_letter(col)].width = 27
ws.column_dimensions["I"].width = 0.5   # separador visual

# Configuración de impresión
ws.page_setup.orientation  = "landscape"
ws.page_setup.paperSize    = 9
ws.page_setup.fitToPage    = True
ws.page_setup.fitToWidth   = 1
ws.page_setup.fitToHeight  = 0
ws.page_margins = PageMargins(left=0.45, right=0.45, top=0.55, bottom=0.55,
                               header=0.2, footer=0.2)
ws.print_title_rows = "1:5"

# ── FILA 1: Encabezado ────────────────────────────────────────────────────────
ws.merge_cells("A1:I1")
ws["A1"] = "CLÍNICA KARINA LARA  ·  Nutrición Clínica Especializada"
ws["A1"].font = Font(name="Arial", bold=True, size=13, color=BLANCO)
ws["A1"].fill = fill(ROJO)
ws["A1"].alignment = ca()
ws.row_dimensions[1].height = 28

# ── FILA 2: Subtítulo ─────────────────────────────────────────────────────────
ws.merge_cells("A2:I2")
ws["A2"] = "MENÚ SEMANAL PERSONALIZADO  ·  Basado en Plan Nutricional Individual  ·  Guía del Plato del Bien Comer"
ws["A2"].font = Font(name="Arial", size=10, italic=True, color="5C1521")
ws["A2"].fill = fill(ROJO_CLR)
ws["A2"].alignment = ca()
ws.row_dimensions[2].height = 18

# ── FILA 3: Datos del paciente ─────────────────────────────────────────────
# Paciente (cols A-C)
ws["A3"] = "Paciente:"
ws["A3"].font = fnt(GRIS_H, b=True, sz=10)
ws["A3"].fill = fill(GRIS_CLR)
ws["A3"].alignment = lc()
ws.merge_cells("B3:C3")
ws["B3"].value = nombre_pac + (f"  ({edad_pac} años)" if edad_pac else "")
ws["B3"].font  = Font(name="Arial", bold=True, size=11, color=ROJO)
ws["B3"].fill  = fill(BLANCO)
ws["B3"].alignment = lc()
ws["B3"].border = Border(bottom=Side(style="thin", color="999999"))

# Semana (cols D-F)
ws["D3"] = "Semana:"
ws["D3"].font = fnt(GRIS_H, b=True, sz=10)
ws["D3"].fill = fill(GRIS_CLR)
ws["D3"].alignment = lc()
ws.merge_cells("E3:F3")
ws["E3"].value = semana_actual
ws["E3"].font  = fnt(sz=10)
ws["E3"].fill  = fill(BLANCO)
ws["E3"].border = Border(bottom=Side(style="thin", color="999999"))

# Nutrióloga (cols G-I)
ws["G3"] = "Nutrióloga:"
ws["G3"].font = fnt(GRIS_H, b=True, sz=10)
ws["G3"].fill = fill(GRIS_CLR)
ws["G3"].alignment = lc()
ws.merge_cells("H3:I3")
ws["H3"].value = "Lic. Karina Lara"
ws["H3"].font  = fnt(sz=10)
ws["H3"].fill  = fill(BLANCO)
ws["H3"].border = Border(bottom=Side(style="thin", color="999999"))
ws.row_dimensions[3].height = 22

# ── FILA 4: Vacía ─────────────────────────────────────────────────────────────
ws.row_dimensions[4].height = 5

# ── FILA 5: Encabezados de días ───────────────────────────────────────────────
ws["A5"] = "TIEMPO\nDE COMIDA"
ws["A5"].font      = Font(name="Arial", bold=True, size=9, color=BLANCO)
ws["A5"].fill      = fill(GRIS_H)
ws["A5"].alignment = ca()
ws["A5"].border    = ba()

for col_idx, dia in enumerate(DIAS, start=2):
    c = ws.cell(row=5, column=col_idx)
    c.value     = dia
    c.font      = Font(name="Arial", bold=True, size=10, color=BLANCO)
    c.fill      = fill(GRIS_H)
    c.alignment = ca()
    c.border    = ba()

# Columna I = separador vacío
ws.cell(row=5, column=9).fill = fill(GRIS_CLR)
ws.row_dimensions[5].height = 32

# ── Bloques de comidas ────────────────────────────────────────────────────────
ROWS_PER_MEAL = 9
current_row = 6

for comida_label, color_h, color_bg, clave in COMIDAS:
    start_r = current_row
    end_r   = current_row + ROWS_PER_MEAL - 1

    # Etiqueta rotada (col A)
    ws.merge_cells(f"A{start_r}:A{end_r}")
    lbl = ws[f"A{start_r}"]
    lbl.value     = comida_label
    lbl.font      = Font(name="Arial", bold=True, size=10, color=BLANCO)
    lbl.fill      = fill(color_h)
    lbl.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True, textRotation=90)
    lbl.border    = bo()

    # Contenido por día (cols B–H)
    for di, dia in enumerate(DIAS):
        col_idx = di + 2
        texto = (menu.get(dia) or {}).get(clave, "—")
        ws.merge_cells(start_row=start_r, start_column=col_idx,
                       end_row=end_r,   end_column=col_idx)
        c = ws.cell(row=start_r, column=col_idx)
        c.value     = texto
        c.font      = Font(name="Arial", size=9, color=NEGRO)
        c.fill      = fill(color_bg)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border    = ba()

    # Separador col I
    ws.merge_cells(start_row=start_r, start_column=9,
                   end_row=end_r,   end_column=9)
    ws.cell(row=start_r, column=9).fill   = fill(GRIS_CLR)
    ws.cell(row=start_r, column=9).border = ba()

    for r in range(start_r, end_r + 1):
        ws.row_dimensions[r].height = 20

    current_row = end_r + 1

# ── Separador ─────────────────────────────────────────────────────────────────
ws.row_dimensions[current_row].height = 5
current_row += 1

# ── Fila HIDRATACIÓN / Nota ───────────────────────────────────────────────────
ws.merge_cells(f"A{current_row}:I{current_row}")
ws[f"A{current_row}"] = (
    "💧 HIDRATACIÓN: Mínimo 1–1.5 L de agua natural al día.  "
    "🍬 Sin azúcar añadida ni bebidas azucaradas.  "
    "🧂 Limitar sal y evitar ultraprocesados.  "
    "⚠️ Adaptar porciones según peso, edad y tolerancia individual."
)
ws[f"A{current_row}"].font = Font(name="Arial", size=8, italic=True, color="5C4033")
ws[f"A{current_row}"].fill = fill(BG_NOTA)
ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws[f"A{current_row}"].border = Border(
    left=Side(style="medium", color=DORADO), right=ba().right,
    top=ba().top, bottom=ba().bottom
)
ws.row_dimensions[current_row].height = 28
current_row += 1

# ── Plato del Bien Comer ──────────────────────────────────────────────────────
ws.merge_cells(f"A{current_row}:I{current_row}")
ws[f"A{current_row}"] = (
    "🍽️  PLATO DEL BIEN COMER:  "
    "½ plato = Verduras y frutas  |  "
    "¼ plato = Cereales y tubérculos integrales  |  "
    "¼ plato = Leguminosas y alimentos de origen animal  |  "
    f"✅ Plan generado el {datetime.now().strftime('%d/%m/%Y')}  ·  Lic. Karina Lara"
)
ws[f"A{current_row}"].font = Font(name="Arial", size=8, color=GRIS_H)
ws[f"A{current_row}"].fill = fill("F0F4F8")
ws[f"A{current_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws[f"A{current_row}"].border = ba()
ws.row_dimensions[current_row].height = 22

# ── Guardar ───────────────────────────────────────────────────────────────────
wb.save(output_path)
print(f"OK: {output_path}", flush=True)
